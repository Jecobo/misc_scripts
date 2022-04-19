# Time   ：2022/4/17 10:40
# Author : Timer Zz
# Email  : 2540373135@qq.com

from core.sql_helper import SQL_helper
from openpyxl import Workbook
import time
from rich.console import Console


def export_data():
    console = Console()
    now_time = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
    wb = Workbook()
    ws_subdomain = wb.create_sheet('subdomains', 0)
    ws_subdomain.append(['ID', 'SUBDOMAIN', 'IP', 'CDN', 'SUBDOMAIN_TIME', 'TAG'])
    ws_urls = wb.create_sheet('urls', 1)
    ws_urls.append(['ID', 'URL', 'isAlive', 'status_code', 'title', 'URL_TIME', 'URL_TAG'])

    subdomains_li = SQL_helper.read_subdomain_sql()
    for subs in subdomains_li:
        ws_subdomain.append(subs)

    urls_li = SQL_helper.url_table_read()
    for urlx in urls_li:
        ws_urls.append(urlx)
    try:
        wb.save('./export/' + now_time + '.xslx')
        console.print("[info] export at ./report/" + now_time + ".xlsx", style='bold blue')
    except:
        console.print("[error] export failed!", style='bold red')


if __name__ == '__main__':
    export_data()
